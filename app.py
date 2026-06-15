import os
import json
import uuid
import sqlite3
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, send_from_directory
from database import get_db, init_db
from docx_generator import generar_docx_alta, generar_docx_baja, generar_docx_prestamo

# ── CONFIGURACIÓN INICIAL ──────────────────────────────────────
app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARCH_DIR = os.path.join(BASE_DIR, 'archivos')
os.makedirs(ARCH_DIR, exist_ok=True)

@app.before_request
def log_request():
    print(f"-> {request.method} {request.path}")

def generar_id():   
    return str(uuid.uuid4())[:8]

def fecha_hoy(): 
    return datetime.now().strftime('%Y-%m-%d')

# ── MIGRACIÓN AUTOMÁTICA DE BASE DE DATOS ──────────────────────
def migrar_db():
    conexion = sqlite3.connect('database.db')
    cursor = conexion.cursor()
    
    # 1. Agregar columnas si no existen
    columnas_nuevas = [
        ("categoria", "TEXT"), ("licencia_windows", "TEXT"), 
        ("version_office", "TEXT"), ("contrasena", "TEXT"), ("numero_tel", "TEXT")
    ]
    for columna, tipo in columnas_nuevas:
        try: cursor.execute(f"ALTER TABLE inventario ADD COLUMN {columna} {tipo}")
        except: pass 
    
    # Agregar columna fecha_asig en empleados y bajas si no existen
    try: cursor.execute("ALTER TABLE empleados ADD COLUMN fecha_asig TEXT")
    except: pass
    try: cursor.execute("ALTER TABLE bajas ADD COLUMN fecha_asig TEXT")
    except: pass
    
    # 2. Permitir IPs duplicadas controladas
    try:
        cursor.execute('''CREATE TABLE IF NOT EXISTS ips_new (
                            id TEXT PRIMARY KEY, direccion TEXT NOT NULL, subred TEXT, 
                            estado TEXT DEFAULT 'libre', asignada_a TEXT, tipo_disp TEXT, 
                            notas TEXT, inv_id TEXT, created_at TEXT DEFAULT CURRENT_TIMESTAMP)''')
        cursor.execute("INSERT OR IGNORE INTO ips_new SELECT id, direccion, subred, estado, asignada_a, tipo_disp, notas, inv_id, created_at FROM ips")
        cursor.execute("DROP TABLE ips")
        cursor.execute("ALTER TABLE ips_new RENAME TO ips")
    except Exception as e:
        print(f"Migración IPs: {e}")
        
    # 3. Tabla para Préstamos
    cursor.execute('''CREATE TABLE IF NOT EXISTS prestamos (
        id TEXT PRIMARY KEY, nombre_solicitante TEXT NOT NULL, area TEXT,
        fecha_prestamo TEXT, fecha_esperada TEXT, estado TEXT DEFAULT 'activo',
        notas TEXT, equipos TEXT DEFAULT '[]', archivo_prestamo TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    # 4. Tabla para Hojas de Servicio
    cursor.execute('''CREATE TABLE IF NOT EXISTS hojas_servicio (
        id             TEXT PRIMARY KEY,
        impresora_id   TEXT NOT NULL,
        nombre_archivo TEXT NOT NULL,
        nombre_visible TEXT NOT NULL,
        fecha          TEXT NOT NULL,
        categoria      TEXT NOT NULL,
        notas          TEXT,
        created_at     TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(impresora_id) REFERENCES impresoras(id) ON DELETE CASCADE
    )''')

    # 5. Sincronizar tabla de impresoras para tener EXACTAMENTE las 21 indicadas
    printers_list = [
        ("R4V2736209", "KYOCERA", "ECOSYS M3145idn", "Archivo Clínico"),
        ("19X3Z08606", "KYOCERA", "TASKalfa 2554ci", "Dirección Médica"),
        ("VR93313757", "KYOCERA", "ECOSYS M2040dn", "Biomédica-2"),
        ("R4P2Y15993", "KYOCERA", "ECOSYS M3655idn", "Médicos Internos"),
        ("REZ3902050", "KYOCERA", "ECOSYS M6230cidn", "Fundación Escandón"),
        ("VR93Y22181", "KYOCERA", "ECOSYS M2040dn", "Dirección Administrativa"),
        ("VR93Y22179", "KYOCERA", "ECOSYS M2040dn", "Caja General"),
        ("11L2Y00092", "KYOCERA", "ECOSYS MA4500ix", "Programación Citas"),
        ("REZ3902091", "KYOCERA", "ECOSYS M6230cidn", "Laboratorio Jefatura"),
        ("REZ3902074", "KYOCERA", "ECOSYS M6230cidn", "Sistemas"),
        ("11L2Y00107", "KYOCERA", "ECOSYS MA4500ix", "Farmacia"),
        ("11L2Y00119", "KYOCERA", "ECOSYS MA4500ix", "Enfermería"),
        ("11L2Y00105", "KYOCERA", "ECOSYS MA4500ix", "Contraloría"),
        ("11L3803760", "KYOCERA", "ECOSYS MA4500ix", "Escandón Almacén"),
        ("11L2Y00109", "KYOCERA", "ECOSYS MA4500ix", "Laboratorio Recepción"),
        ("11L3803759", "KYOCERA", "ECOSYS MA4500ix", "Escuela Enfermería"),
        ("R4V2Z47074", "KYOCERA", "ECOSYS M3145idn", "Recursos Humanos"),
        ("W6A3702250", "KYOCERA", "ECOSYS MA2100cwfx", "Dirección General / Contador"),
        ("11L3803867", "KYOCERA", "ECOSYS MA4500ix", "Trabajo Social"),
        ("11L2Y00106", "KYOCERA", "ECOSYS MA4500ix", "Calidad"),
        ("WDS4Z53505", "KYOCERA", "ECOSYS MA5500ifx", "Análisis de Cuentas")
    ]
    
    allowed_serials = {p[0] for p in printers_list}
    
    # Obtener todas las impresoras actuales en la BD
    actuales = cursor.execute("SELECT id, nombre, ip_id FROM impresoras").fetchall()
    
    # 1. Eliminar las que no están en la lista permitida
    for imp_id, nombre, ip_id in actuales:
        if nombre not in allowed_serials:
            # Desvincular IP
            if ip_id:
                cursor.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, asignada_a=NULL, inv_id=NULL WHERE id=?", (ip_id,))
            # Eliminar vinculaciones de equipos
            cursor.execute("DELETE FROM impresora_equipos WHERE imp_id=?", (imp_id,))
            # Eliminar la impresora
            cursor.execute("DELETE FROM impresoras WHERE id=?", (imp_id,))
            
    # 2. Agregar o actualizar las de la lista
    for serial, marca, modelo, ubicacion in printers_list:
        row = cursor.execute("SELECT id FROM impresoras WHERE nombre=?", (serial,)).fetchone()
        if row:
            cursor.execute("UPDATE impresoras SET marca=?, modelo=?, ubicacion=? WHERE nombre=?", 
                           (marca, modelo, ubicacion, serial))
        else:
            imp_id = "imp_" + serial.lower()
            cursor.execute('''INSERT INTO impresoras 
                              (id, nombre, marca, modelo, tipo_consumible, modelo_consumible, ubicacion, ip_id, conexion, estado, notas)
                              VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                           (imp_id, serial, marca, modelo, "Tóner", "", ubicacion, None, "red", "activa", "Arrendada"))
        
    conexion.commit()
    conexion.close()

# ── HELPERS ────────────────────────────────────────────────────
def obtener_equipos_de_empleado(empleado_id):
    db = get_db()
    filas = db.execute('''
        SELECT a.id as asig_id, a.cantidad, a.fecha_asig,
               i.id as inv_id, i.tipo, i.marca, i.modelo,
               i.num_serie, i.procesador, i.ram, i.almacenamiento,
               i.licencia_windows, i.version_office, i.contrasena, i.numero_tel,
               ip.direccion as ip_asignada
        FROM asignaciones a
        JOIN inventario i ON a.inv_id = i.id
        LEFT JOIN ips ip ON i.ip_id = ip.id
        WHERE a.emp_id = ? AND a.activa = 1
    ''', (empleado_id,)).fetchall()
    db.close()
    return [dict(fila) for fila in filas]

def obtener_usuarios_de_equipo(inventario_id):
    db = get_db()
    filas = db.execute('''
        SELECT a.cantidad, e.id as emp_id, e.nombre, e.area
        FROM asignaciones a 
        JOIN empleados e ON a.emp_id = e.id
        WHERE a.inv_id = ? AND a.activa = 1 AND e.activo = 1
    ''', (inventario_id,)).fetchall()
    db.close()
    return [dict(fila) for fila in filas]

def ordenar_ips(lista_ips):
    def tupla_ip(fila):
        try: return tuple(int(x) for x in dict(fila)['direccion'].split('.'))
        except ValueError: return (0,0,0,0)
    return sorted(lista_ips, key=tupla_ip)

# ── RUTAS DE VISTAS ────────────────────────────────────────────
@app.route('/')
def index(): 
    return render_template('index.html')

@app.route('/static/<path:archivo>')
def servir_estatico(archivo): 
    return send_from_directory(os.path.join(BASE_DIR, 'static'), archivo)

# ── RUTAS DE API: ESTADÍSTICAS ─────────────────────────────────
@app.route('/api/stats')
def api_stats():
    try:
        db = get_db()
        activos = db.execute('SELECT COUNT(*) FROM empleados WHERE activo=1').fetchone()[0]
        bajas = db.execute('SELECT COUNT(*) FROM bajas').fetchone()[0]
        prestamos_activos = db.execute("SELECT COUNT(*) FROM prestamos WHERE estado='activo'").fetchone()[0]
        inv_total = db.execute('SELECT COUNT(*) FROM inventario').fetchone()[0]
        inv_libre = db.execute("SELECT SUM(cantidad_libre) FROM inventario").fetchone()[0] or 0
        ips_libre = db.execute("SELECT COUNT(*) FROM ips WHERE estado='libre'").fetchone()[0]
        ips_asig = db.execute("SELECT COUNT(*) FROM ips WHERE estado='asignada'").fetchone()[0]
        ips_duplicadas = db.execute('''SELECT COUNT(*) FROM (SELECT direccion FROM ips WHERE estado='asignada' AND tipo_disp != 'Cámara' GROUP BY direccion HAVING COUNT(*)>1)''').fetchone()[0]
        imp_total = db.execute('SELECT COUNT(*) FROM impresoras').fetchone()[0]
        imp_alertas = db.execute('''SELECT COUNT(*) FROM (SELECT imp_id FROM impresora_equipos GROUP BY imp_id HAVING COUNT(*)>5)''').fetchone()[0]
        cam_total = db.execute('SELECT COUNT(*) FROM camaras').fetchone()[0]
        cam_sin_ip = db.execute('SELECT COUNT(*) FROM camaras WHERE ip_id IS NULL').fetchone()[0]
        red_total = db.execute('SELECT COUNT(*) FROM red_infra').fetchone()[0]
        inv_sin_asig = db.execute("SELECT SUM(cantidad_libre) FROM inventario WHERE cantidad_libre=cantidad_total AND cantidad_total>0").fetchone()[0] or 0
        db.close()
        
        return jsonify({
            "activos": activos, "bajas": bajas, "prestamos_activos": prestamos_activos, "inv_tot": inv_total, "inv_libre": inv_libre,
            "ips_libre": ips_libre, "ips_asig": ips_asig, "ips_dup": ips_duplicadas,
            "imp_tot": imp_total, "imp_alert": imp_alertas, "cam_tot": cam_total, 
            "cam_sin_ip": cam_sin_ip, "red_tot": red_total, "inv_sin_asig": inv_sin_asig
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── RUTAS DE API: EMPLEADOS ────────────────────────────────────
@app.route('/api/empleados')
def listar_empleados():
    db = get_db()
    filas = db.execute('SELECT * FROM empleados WHERE activo=1 ORDER BY nombre').fetchall()
    db.close()
    resultado = []
    for fila in filas:
        empleado = dict(fila)
        empleado['cuentas'] = json.loads(empleado['cuentas'] or '{}')
        empleado['equipos'] = obtener_equipos_de_empleado(empleado['id'])
        resultado.append(empleado)
    return jsonify(resultado)

@app.route('/api/empleados/<empleado_id>')
def detalle_empleado(empleado_id):
    db = get_db()
    fila = db.execute('SELECT * FROM empleados WHERE id=?', (empleado_id,)).fetchone()
    db.close()
    if not fila: return jsonify({'error': 'Empleado no encontrado'}), 404
    empleado = dict(fila)
    empleado['cuentas'] = json.loads(empleado['cuentas'] or '{}')
    empleado['equipos'] = obtener_equipos_de_empleado(empleado_id)
    return jsonify(empleado)

@app.route('/api/empleados', methods=['POST'])
def crear_empleado():
    try:
        datos = request.json
        empleado_id = generar_id()
        db = get_db()
        db.execute('''INSERT INTO empleados
            (id, nombre, fecha_alta, fecha_asig, direccion, area, ubicacion, telefono, observaciones, cuentas)
            VALUES(?,?,?,?,?,?,?,?,?,?)''',
            (empleado_id, datos['nombre'], datos.get('fechaAlta', fecha_hoy()), fecha_hoy(), datos.get('direccion'),
             datos.get('area'), datos.get('ubicacion'), datos.get('telefono'),
             datos.get('observaciones'), json.dumps(datos.get('cuentas', {}))))
        
        for equipo in datos.get('equipos', []):
            db.execute('INSERT INTO asignaciones(id, emp_id, inv_id, cantidad, fecha_asig, activa) VALUES(?,?,?,?,?,1)',
                       (generar_id(), empleado_id, equipo['inv_id'], equipo.get('cantidad', 1), fecha_hoy()))
            db.execute('UPDATE inventario SET cantidad_libre = cantidad_libre - ? WHERE id = ?', (equipo.get('cantidad', 1), equipo['inv_id']))
            db.execute("UPDATE inventario SET estado = 'prestado' WHERE id = ? AND cantidad_libre <= 0", (equipo['inv_id'],))
        
        db.commit(); db.close()
        return jsonify({'id': empleado_id, 'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/empleados/<empleado_id>', methods=['PUT'])
def actualizar_empleado(empleado_id):
    try:
        datos = request.json
        db = get_db()
        db.execute('''UPDATE empleados SET nombre=?, fecha_alta=?, direccion=?, area=?,
                      ubicacion=?, telefono=?, observaciones=?, cuentas=?, fecha_asig=? WHERE id=?''',
                   (datos['nombre'], datos.get('fechaAlta'), datos.get('direccion'), datos.get('area'),
                    datos.get('ubicacion'), datos.get('telefono'), datos.get('observaciones'),
                    json.dumps(datos.get('cuentas', {})), fecha_hoy(), empleado_id))
        db.commit(); db.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/empleados/<empleado_id>/asignaciones', methods=['POST'])
def actualizar_asignaciones(empleado_id):
    try:
        datos = request.json
        db = get_db()
        
        # 1. Liberar lo que el empleado ya tenía asignado
        asignaciones_actuales = db.execute('SELECT * FROM asignaciones WHERE emp_id=? AND activa=1', (empleado_id,)).fetchall()
        for asig in asignaciones_actuales:
            db.execute('UPDATE inventario SET cantidad_libre = cantidad_libre + ? WHERE id = ?', (asig['cantidad'], asig['inv_id']))
            db.execute("UPDATE inventario SET estado = CASE WHEN cantidad_libre >= cantidad_total THEN 'libre' ELSE estado END WHERE id = ?", (asig['inv_id'],))
            db.execute('UPDATE asignaciones SET activa = 0, fecha_devol = ? WHERE id = ?', (fecha_hoy(), asig['id']))
        
        # 2. Asignar los nuevos equipos seleccionados
        for equipo in datos.get('equipos', []):
            db.execute('INSERT INTO asignaciones(id, emp_id, inv_id, cantidad, fecha_asig, activa) VALUES(?,?,?,?,?,1)',
                       (generar_id(), empleado_id, equipo['inv_id'], equipo.get('cantidad', 1), fecha_hoy()))
            db.execute('UPDATE inventario SET cantidad_libre = cantidad_libre - ? WHERE id = ?', (equipo.get('cantidad', 1), equipo['inv_id']))
            db.execute("UPDATE inventario SET estado = 'prestado' WHERE id = ? AND cantidad_libre <= 0", (equipo['inv_id'],))
        
        if 'cuentas' in datos:
            db.execute('UPDATE empleados SET cuentas=? WHERE id=?', (json.dumps(datos['cuentas']), empleado_id))
            
        db.execute('UPDATE empleados SET fecha_asig=? WHERE id=?', (fecha_hoy(), empleado_id))
        db.commit(); db.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── RUTAS DE API: BAJAS ────────────────────────────────────────
@app.route('/api/bajas')
def listar_bajas():
    db = get_db()
    filas = db.execute('SELECT * FROM bajas ORDER BY fecha_baja DESC').fetchall()
    db.close()
    resultado = []
    for fila in filas:
        baja = dict(fila)
        baja['cuentas'] = json.loads(baja['cuentas'] or '{}')
        baja['equipos_snap']= json.loads(baja['equipos_snap'] or '[]')
        resultado.append(baja)
    return jsonify(resultado)

@app.route('/api/bajas', methods=['POST'])
def registrar_baja():
    try:
        datos = request.json
        empleado_id = datos['empId']
        db = get_db()
        empleado = db.execute('SELECT * FROM empleados WHERE id=?', (empleado_id,)).fetchone()
        if not empleado: db.close(); return jsonify({'error': 'Empleado no encontrado'}), 404
            
        equipos = obtener_equipos_de_empleado(empleado_id)
        baja_id = generar_id()
        
        db.execute('''INSERT INTO bajas
            (id, emp_id, nombre, fecha_alta, fecha_asig, fecha_baja, motivo, direccion, area,
             ubicacion, telefono, cuentas, equipos_snap, archivo_alta)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (baja_id, empleado_id, empleado['nombre'], empleado['fecha_alta'], empleado['fecha_asig'], datos.get('fechaBaja', fecha_hoy()),
             datos.get('motivo'), empleado['direccion'], empleado['area'], empleado['ubicacion'],
             empleado['telefono'], empleado['cuentas'], json.dumps(equipos), empleado['archivo_alta']))
             
        for asig in db.execute('SELECT * FROM asignaciones WHERE emp_id=? AND activa=1', (empleado_id,)).fetchall():
            db.execute('UPDATE inventario SET cantidad_libre = cantidad_libre + ? WHERE id = ?', (asig['cantidad'], asig['inv_id']))
            db.execute('''UPDATE inventario SET estado=
                          CASE WHEN cantidad_libre>=cantidad_total THEN 'libre' ELSE estado END
                          WHERE id=?''',(asig['inv_id'],))
            db.execute('UPDATE asignaciones SET activa=0, fecha_devol=? WHERE id=?', (datos.get('fechaBaja', fecha_hoy()), asig['id']))
            
        db.execute('UPDATE empleados SET activo=0 WHERE id=?', (empleado_id,))
        db.commit(); db.close()
        return jsonify({'id': baja_id, 'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── RUTAS DE API: PRÉSTAMOS TEMPORALES ─────────────────────────
@app.route('/api/prestamos')
def listar_prestamos():
    db = get_db()
    filas = db.execute('SELECT * FROM prestamos ORDER BY estado ASC, fecha_prestamo DESC').fetchall()
    db.close()
    resultado = []
    for fila in filas:
        prestamo = dict(fila)
        prestamo['equipos'] = json.loads(prestamo['equipos'] or '[]')
        resultado.append(prestamo)
    return jsonify(resultado)

@app.route('/api/prestamos', methods=['POST'])
def crear_prestamo():
    try:
        datos = request.json
        pid = generar_id()
        db = get_db()
        db.execute('''INSERT INTO prestamos
            (id, nombre_solicitante, area, fecha_prestamo, fecha_esperada, estado, notas, equipos)
            VALUES(?,?,?,?,?,'activo',?,?)''',
            (pid, datos['nombre_solicitante'], datos.get('area'), datos.get('fecha_prestamo', fecha_hoy()),
             datos.get('fecha_esperada'), datos.get('notas'), json.dumps(datos.get('equipos', []))))
        
        for eq in datos.get('equipos', []):
            db.execute('UPDATE inventario SET cantidad_libre = cantidad_libre - ? WHERE id = ?', (eq.get('cantidad', 1), eq['inv_id']))
            db.execute("UPDATE inventario SET estado = 'prestado' WHERE id = ? AND cantidad_libre <= 0", (eq['inv_id'],))
            
        db.commit(); db.close()
        return jsonify({'id': pid, 'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/prestamos/<pid>/devolver', methods=['PUT'])
def devolver_prestamo(pid):
    try:
        db = get_db()
        p = db.execute('SELECT * FROM prestamos WHERE id=?', (pid,)).fetchone()
        if not p: return jsonify({'error':'Préstamo no encontrado'}), 404
        
        if p['estado'] != 'devuelto':
            equipos = json.loads(p['equipos'] or '[]')
            for eq in equipos:
                db.execute('UPDATE inventario SET cantidad_libre = cantidad_libre + ? WHERE id = ?', (eq.get('cantidad', 1), eq['inv_id']))
                db.execute("UPDATE inventario SET estado = CASE WHEN cantidad_libre >= cantidad_total THEN 'libre' ELSE estado END WHERE id = ?", (eq['inv_id'],))
                
            db.execute("UPDATE prestamos SET estado='devuelto' WHERE id=?", (pid,))
            db.commit()
        db.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── RUTAS DE API: INVENTARIO ───────────────────────────────────
CATEGORIA_MAP = {
    'PC': 'Cómputo', 'Laptop': 'Cómputo', 'Monitor': 'Cómputo',
    'Mouse': 'Periféricos', 'Teclado': 'Periféricos', 'Audífonos Diadema': 'Periféricos',
    'Adaptador RED-USB': 'Periféricos', 'Memoria USB': 'Periféricos', 'Cargador': 'Periféricos',
    'Lector de código de barras': 'Periféricos', 'Proyector': 'Audiovisual',
    'Impresora de Tinta': 'Impresión', 'Impresora de Cartucho': 'Impresión', 
    'Impresora de Tóner': 'Impresión', 'Impresora Térmica': 'Impresión', 'Impresora de etiquetas': 'Impresión',
    'Teléfono IP': 'Telefonía', 'Teléfono Conmutador': 'Telefonía', 'Radio': 'Telefonía', 'Celular': 'Telefonía',
    'No Break': 'Energía', 'Servidor': 'Red', 'Switch': 'Red', 'Router': 'Red', 'Access Point': 'Red',
    'Patch Panel': 'Red', 'Rack': 'Red', 'UPS': 'Energía', 'Firewall': 'Red', 'KVM': 'Red',
    'Smart TV': 'Audiovisual'
}

@app.route('/api/inventario')
def listar_inventario():
    db = get_db()
    filas = db.execute('''
        SELECT i.*, ip.direccion as ip_dir
        FROM inventario i
        LEFT JOIN ips ip ON i.ip_id = ip.id
        ORDER BY i.tipo, i.marca, i.modelo
    ''').fetchall()
    db.close()
    resultado = []
    for fila in filas:
        item = dict(fila)
        item['asignados_a'] = obtener_usuarios_de_equipo(item['id'])
        resultado.append(item)
    return jsonify(resultado)

@app.route('/api/inventario/libres')
def listar_inventario_libre():
    db = get_db()
    filas = db.execute('SELECT i.*, ip.direccion as ip_dir FROM inventario i LEFT JOIN ips ip ON i.ip_id = ip.id WHERE i.cantidad_libre > 0 ORDER BY i.tipo, i.marca, i.modelo').fetchall()
    db.close()
    return jsonify([dict(fila) for fila in filas])

@app.route('/api/inventario/<inv_id>')
def get_inv_item(inv_id):
    db = get_db()
    r = db.execute('SELECT i.*,ip.direccion as ip_dir FROM inventario i LEFT JOIN ips ip ON i.ip_id=ip.id WHERE i.id=?',(inv_id,)).fetchone()
    db.close()
    if not r: return jsonify({'error':'no encontrado'}),404
    item = dict(r); item['asignados_a']=obtener_usuarios_de_equipo(inv_id)
    return jsonify(item)

@app.route('/api/inventario', methods=['POST'])
def crear_inventario():
    try:
        datos = request.json
        inv_id = generar_id()
        cantidad = int(datos.get('cantidad_total', 1))
        categoria = CATEGORIA_MAP.get(datos['tipo'], 'General')
        
        db = get_db()
        db.execute('''INSERT INTO inventario
            (id, tipo, categoria, marca, modelo, num_serie, ip_id, procesador, ram, almacenamiento,
             licencia_windows, version_office, contrasena, numero_tel, cantidad_total, cantidad_libre, estado, notas)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (inv_id, datos['tipo'], categoria, datos.get('marca'), datos.get('modelo'), datos.get('num_serie'),
             datos.get('ip_id'), datos.get('procesador'), datos.get('ram'), datos.get('almacenamiento'),
             datos.get('licencia_windows'), datos.get('version_office'), datos.get('contrasena'), datos.get('numero_tel'),
             cantidad, cantidad, 'libre', datos.get('notas')))
             
        if datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='asignada', tipo_disp=?, inv_id=? WHERE id=?", (datos['tipo'], inv_id, datos['ip_id']))
        db.commit(); db.close()
        return jsonify({'id': inv_id, 'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/inventario/<inv_id>', methods=['PUT'])
def actualizar_inventario(inv_id):
    try:
        datos = request.json
        categoria = CATEGORIA_MAP.get(datos.get('tipo', ''), 'General')
        db = get_db()
        previo = db.execute('SELECT ip_id, cantidad_total, cantidad_libre FROM inventario WHERE id=?', (inv_id,)).fetchone()
        
        if previo['ip_id'] and previo['ip_id'] != datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, inv_id=NULL WHERE id=?", (previo['ip_id'],))
        
        total_anterior = previo['cantidad_total']
        libre_anterior = previo['cantidad_libre']
        nuevo_total = int(datos.get('cantidad_total', 1))
        nuevo_libre = libre_anterior + (nuevo_total - total_anterior)
        if nuevo_libre < 0: nuevo_libre = 0 

        db.execute('''UPDATE inventario SET tipo=?, categoria=?, marca=?, modelo=?, num_serie=?, ip_id=?,
                      procesador=?, ram=?, almacenamiento=?, licencia_windows=?, version_office=?,
                      contrasena=?, numero_tel=?, cantidad_total=?, cantidad_libre=?, notas=?, estado=? WHERE id=?''',
                   (datos['tipo'], categoria, datos.get('marca'), datos.get('modelo'), datos.get('num_serie'), datos.get('ip_id'),
                    datos.get('procesador'), datos.get('ram'), datos.get('almacenamiento'),
                    datos.get('licencia_windows'), datos.get('version_office'), datos.get('contrasena'), datos.get('numero_tel'),
                    nuevo_total, nuevo_libre, datos.get('notas'), datos.get('estado', 'libre'), inv_id))
        
        if datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='asignada', tipo_disp=?, inv_id=? WHERE id=?", (datos['tipo'], inv_id, datos['ip_id']))
        db.commit(); db.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/inventario/<inv_id>', methods=['DELETE'])
def eliminar_inventario(inv_id):
    db = get_db()
    asignaciones_activas = db.execute('SELECT COUNT(*) FROM asignaciones WHERE inv_id=? AND activa=1', (inv_id,)).fetchone()[0]
    if asignaciones_activas > 0: 
        db.close(); return jsonify({'error': 'El equipo tiene asignaciones activas'}), 400
    
    db.execute('DELETE FROM asignaciones WHERE inv_id=?', (inv_id,))
    db.execute('DELETE FROM impresora_equipos WHERE inv_id=?', (inv_id,))
    db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, inv_id=NULL WHERE inv_id=?", (inv_id,))
    db.execute('DELETE FROM inventario WHERE id=?', (inv_id,))
    db.commit(); db.close()
    return jsonify({'ok': True})

# ── RUTAS DE API: IPs ──────────────────────────────────────────
def enriquecer_datos_ip(fila_ip, db):
    ip = dict(fila_ip)
    ip['link'] = None
    if ip.get('estado') != 'asignada': return ip
    
    if ip.get('inv_id'):
        rel = db.execute('''SELECT i.id, i.tipo, i.marca, i.modelo, a.emp_id, e.nombre as emp_nombre FROM inventario i
            LEFT JOIN asignaciones a ON a.inv_id = i.id AND a.activa = 1
            LEFT JOIN empleados e ON e.id = a.emp_id AND e.activo = 1 
            WHERE i.id = ?''', (ip['inv_id'],)).fetchone()
        if rel:
            rel = dict(rel)
            if rel.get('emp_id'):
                ip['link'] = {'tipo': 'empleado', 'id': rel['emp_id'], 'nombre': rel['emp_nombre']}
                ip['asignada_a'] = rel['emp_nombre']
            ip['tipo_disp'] = f"{rel['tipo']} {rel.get('marca', '')} {rel.get('modelo', '')}".strip()
    
    imp = db.execute("SELECT id, nombre, marca, modelo FROM impresoras WHERE ip_id=?", (ip['id'],)).fetchone()
    if imp:
        imp = dict(imp)
        ip['link'] = {'tipo': 'impresora', 'id': imp['id'], 'nombre': imp['nombre'] or f"{imp['marca']} {imp['modelo']}"}
        ip['tipo_disp'] = 'Impresora'
        if not ip.get('asignada_a'): ip['asignada_a'] = ip['link']['nombre']

    cam = db.execute("SELECT id, area, ubicacion FROM camaras WHERE ip_id=?", (ip['id'],)).fetchone()
    if cam:
        cam = dict(cam)
        ip['link'] = {'tipo': 'camara', 'id': cam['id'], 'nombre': f"Cámara — {cam['area']}, {cam['ubicacion']}"}
        ip['tipo_disp'] = 'Cámara'
        if not ip.get('asignada_a'): ip['asignada_a'] = ip['link']['nombre']

    red = db.execute("SELECT id, tipo, nombre, marca, modelo FROM red_infra WHERE ip_id=?", (ip['id'],)).fetchone()
    if red:
        red = dict(red)
        ip['link'] = {'tipo': 'red', 'id': red['id'], 'nombre': red['nombre'] or f"{red['tipo']} {red.get('marca', '')}"}
        ip['tipo_disp'] = red['tipo']
        if not ip.get('asignada_a'): ip['asignada_a'] = ip['link']['nombre']
    
    return ip

@app.route('/api/ips')
def listar_ips():
    db = get_db()
    filas = db.execute('SELECT * FROM ips').fetchall()
    ips_ordenadas = ordenar_ips(filas)
    resultado = [enriquecer_datos_ip(fila, db) for fila in ips_ordenadas]
    db.close()
    return jsonify(resultado)

@app.route('/api/ips/libres')
def listar_ips_libres():
    db = get_db()
    filas = db.execute("SELECT * FROM ips WHERE estado='libre'").fetchall()
    db.close()
    return jsonify([dict(fila) for fila in ordenar_ips(filas)])

@app.route('/api/ips/<ip_id>', methods=['PUT'])
def actualizar_ip(ip_id):
    datos = request.json
    db = get_db()
    db.execute('UPDATE ips SET estado=?, asignada_a=?, tipo_disp=?, notas=? WHERE id=?',
               (datos.get('estado', 'libre'), datos.get('asignada_a'), datos.get('tipo_disp'), datos.get('notas'), ip_id))
    db.commit()
    db.close()
    return jsonify({'ok': True})

# ── RUTAS DE API: IMPRESORAS ───────────────────────────────────
@app.route('/api/impresoras')
def listar_impresoras():
    db = get_db()
    filas = db.execute('SELECT i.*, ip.direccion as ip_dir FROM impresoras i LEFT JOIN ips ip ON i.ip_id = ip.id ORDER BY i.ubicacion, i.marca').fetchall()
    resultado = []
    for fila in filas:
        imp = dict(fila)
        equipos = db.execute('''SELECT ie.id, ie.tipo_conn, inv.id as inv_id, inv.tipo, inv.marca, inv.modelo, inv.num_serie 
                                FROM impresora_equipos ie 
                                JOIN inventario inv ON ie.inv_id = inv.id 
                                WHERE ie.imp_id = ?''', (imp['id'],)).fetchall()
        
        lista_equipos = []
        for eq in equipos:
            eq_dict = dict(eq)
            asig = db.execute('SELECT e.nombre, e.area FROM asignaciones a JOIN empleados e ON a.emp_id = e.id WHERE a.inv_id = ? AND a.activa = 1', (eq_dict['inv_id'],)).fetchone()
            eq_dict['asignado_a'] = asig['nombre'] if asig else ''
            eq_dict['asignado_area'] = asig['area'] if asig else ''
            lista_equipos.append(eq_dict)
            
        imp['equipos_conectados'] = lista_equipos
        imp['alerta_exceso'] = len(imp['equipos_conectados']) > 5
        resultado.append(imp)
    db.close()
    return jsonify(resultado)

@app.route('/api/impresoras', methods=['POST'])
def crear_impresora():
    try:
        datos = request.json
        imp_id = generar_id()
        db = get_db()
        db.execute('''INSERT INTO impresoras
            (id, nombre, marca, modelo, tipo_consumible, modelo_consumible, ubicacion, ip_id, conexion, estado, notas)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
            (imp_id, datos.get('nombre'), datos.get('marca'), datos.get('modelo'), datos.get('tipo_consumible'),
             datos.get('modelo_consumible'), datos.get('ubicacion'), datos.get('ip_id'),
             datos.get('conexion', 'red'), datos.get('estado', 'activa'), datos.get('notas')))
             
        if datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='asignada', tipo_disp='Impresora' WHERE id=?", (datos['ip_id'],))
            
        for equipo in datos.get('equipos', []):
            db.execute('INSERT INTO impresora_equipos(id, imp_id, inv_id, tipo_conn) VALUES(?,?,?,?)',
                       (generar_id(), imp_id, equipo['inv_id'], equipo.get('tipo_conn', 'red')))
        db.commit()
        db.close()
        return jsonify({'id': imp_id, 'ok': True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/impresoras/<imp_id>', methods=['PUT'])
def actualizar_impresora(imp_id):
    datos = request.json
    db = get_db()
    previo = db.execute('SELECT ip_id FROM impresoras WHERE id=?', (imp_id,)).fetchone()
    
    if previo and previo['ip_id'] and previo['ip_id'] != datos.get('ip_id'):
        db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, asignada_a=NULL WHERE id=?", (previo['ip_id'],))
        
    db.execute('''UPDATE impresoras SET nombre=?, marca=?, modelo=?, tipo_consumible=?,
                  modelo_consumible=?, ubicacion=?, ip_id=?, conexion=?, estado=?, notas=? WHERE id=?''',
               (datos.get('nombre'), datos.get('marca'), datos.get('modelo'), datos.get('tipo_consumible'),
                datos.get('modelo_consumible'), datos.get('ubicacion'), datos.get('ip_id'),
                datos.get('conexion', 'red'), datos.get('estado', 'activa'), datos.get('notas'), imp_id))
    
    if datos.get('ip_id'):
        db.execute("UPDATE ips SET estado='asignada', tipo_disp='Impresora' WHERE id=?", (datos['ip_id'],))
        
    db.execute('DELETE FROM impresora_equipos WHERE imp_id=?', (imp_id,))
    for equipo in datos.get('equipos', []):
        db.execute('INSERT INTO impresora_equipos(id, imp_id, inv_id, tipo_conn) VALUES(?,?,?,?)',
                   (generar_id(), imp_id, equipo['inv_id'], equipo.get('tipo_conn', 'red')))
    db.commit()
    db.close()
    return jsonify({'ok': True})

@app.route('/api/impresoras/<imp_id>', methods=['DELETE'])
def eliminar_impresora(imp_id):
    db = get_db()
    previo = db.execute('SELECT ip_id FROM impresoras WHERE id=?', (imp_id,)).fetchone()
    if previo and previo['ip_id']:
        db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, asignada_a=NULL WHERE id=?", (previo['ip_id'],))
    
    db.execute('DELETE FROM impresora_equipos WHERE imp_id=?', (imp_id,))
    db.execute('DELETE FROM impresoras WHERE id=?', (imp_id,))
    db.commit()
    db.close()
    return jsonify({'ok': True})

# ── RUTAS DE API: CÁMARAS Y RED ────────────────────────────────
@app.route('/api/camaras', methods=['GET', 'POST'])
def handle_camaras():
    db = get_db()
    if request.method == 'GET':
        filas = db.execute('SELECT c.*, ip.direccion as ip_dir FROM camaras c LEFT JOIN ips ip ON c.ip_id = ip.id ORDER BY c.area, c.ubicacion').fetchall()
        db.close()
        return jsonify([dict(fila) for fila in filas])
        
    if request.method == 'POST':
        datos = request.json
        cam_id = generar_id()
        db.execute('INSERT INTO camaras(id, area, ubicacion, marca, modelo, ip_id, estado, notas) VALUES(?,?,?,?,?,?,?,?)',
                   (cam_id, datos['area'], datos['ubicacion'], datos.get('marca'), datos.get('modelo'), datos.get('ip_id'), datos.get('estado', 'activa'), datos.get('notas')))
        if datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='asignada', tipo_disp='Cámara' WHERE id=?", (datos['ip_id'],))
        db.commit()
        db.close()
        return jsonify({'id': cam_id, 'ok': True})

@app.route('/api/camaras/<cam_id>', methods=['PUT', 'DELETE'])
def handle_camara_item(cam_id):
    db = get_db()
    if request.method == 'PUT':
        datos = request.json
        previo = db.execute('SELECT ip_id FROM camaras WHERE id=?', (cam_id,)).fetchone()
        if previo and previo['ip_id'] and previo['ip_id'] != datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, asignada_a=NULL WHERE id=?", (previo['ip_id'],))
        db.execute('UPDATE camaras SET area=?, ubicacion=?, marca=?, modelo=?, ip_id=?, estado=?, notas=? WHERE id=?',
                   (datos['area'], datos['ubicacion'], datos.get('marca'), datos.get('modelo'), datos.get('ip_id'), datos.get('estado', 'activa'), datos.get('notas'), cam_id))
        if datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='asignada', tipo_disp='Cámara' WHERE id=?", (datos['ip_id'],))
        db.commit()
        db.close()
        return jsonify({'ok': True})
        
    if request.method == 'DELETE':
        previo = db.execute('SELECT ip_id FROM camaras WHERE id=?', (cam_id,)).fetchone()
        if previo and previo['ip_id']:
            db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, asignada_a=NULL WHERE id=?", (previo['ip_id'],))
        db.execute('DELETE FROM camaras WHERE id=?', (cam_id,))
        db.commit()
        db.close()
        return jsonify({'ok': True})

@app.route('/api/red', methods=['GET', 'POST'])
def handle_red():
    db = get_db()
    if request.method == 'GET':
        filas = db.execute('SELECT r.*, ip.direccion as ip_dir FROM red_infra r LEFT JOIN ips ip ON r.ip_id = ip.id ORDER BY r.tipo, r.nombre').fetchall()
        db.close()
        return jsonify([dict(fila) for fila in filas])
        
    if request.method == 'POST':
        datos = request.json
        red_id = generar_id()
        db.execute('''INSERT INTO red_infra(id, tipo, nombre, marca, modelo, num_serie, ip_id, ubicacion, puertos, notas, estado)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
                   (red_id, datos['tipo'], datos.get('nombre'), datos.get('marca'), datos.get('modelo'), datos.get('num_serie'),
                    datos.get('ip_id'), datos.get('ubicacion'), datos.get('puertos'), datos.get('notas'), datos.get('estado', 'activo')))
        if datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='asignada', tipo_disp=? WHERE id=?", (datos['tipo'], datos['ip_id']))
        db.commit()
        db.close()
        return jsonify({'id': red_id, 'ok': True})

@app.route('/api/red/<red_id>', methods=['PUT', 'DELETE'])
def handle_red_item(red_id):
    db = get_db()
    if request.method == 'PUT':
        datos = request.json
        previo = db.execute('SELECT ip_id FROM red_infra WHERE id=?', (red_id,)).fetchone()
        if previo and previo['ip_id'] and previo['ip_id'] != datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, asignada_a=NULL WHERE id=?", (previo['ip_id'],))
        db.execute('''UPDATE red_infra SET tipo=?, nombre=?, marca=?, modelo=?, num_serie=?,
                      ip_id=?, ubicacion=?, puertos=?, notas=?, estado=? WHERE id=?''',
                   (datos['tipo'], datos.get('nombre'), datos.get('marca'), datos.get('modelo'), datos.get('num_serie'),
                    datos.get('ip_id'), datos.get('ubicacion'), datos.get('puertos'), datos.get('notas'), datos.get('estado', 'activo'), red_id))
        if datos.get('ip_id'):
            db.execute("UPDATE ips SET estado='asignada', tipo_disp=? WHERE id=?", (datos['tipo'], datos['ip_id']))
        db.commit()
        db.close()
        return jsonify({'ok': True})
        
    if request.method == 'DELETE':
        previo = db.execute('SELECT ip_id FROM red_infra WHERE id=?', (red_id,)).fetchone()
        if previo and previo['ip_id']:
            db.execute("UPDATE ips SET estado='libre', tipo_disp=NULL, asignada_a=NULL WHERE id=?", (previo['ip_id'],))
        db.execute('DELETE FROM red_infra WHERE id=?', (red_id,))
        db.commit()
        db.close()
        return jsonify({'ok': True})

# ── RUTAS DE API: HOJAS DE SERVICIO (MANTENIMIENTO) ────────────
@app.route('/api/hojas_servicio', methods=['GET', 'POST'])
def handle_hojas_servicio():
    db = get_db()
    if request.method == 'GET':
        filas = db.execute('''
            SELECT hs.*, imp.nombre as imp_serial, imp.marca as imp_marca, imp.modelo as imp_modelo, imp.ubicacion as imp_ubicacion
            FROM hojas_servicio hs
            JOIN impresoras imp ON hs.impresora_id = imp.id
            ORDER BY hs.fecha DESC, hs.created_at DESC
        ''').fetchall()
        db.close()
        return jsonify([dict(fila) for fila in filas])
        
    if request.method == 'POST':
        try:
            archivo = request.files.get('archivo')
            if not archivo or archivo.filename == '':
                db.close()
                return jsonify({'error': 'Selecciona un archivo PDF válido'}), 400
                
            impresora_id = request.form.get('impresora_id')
            nombre_visible = request.form.get('nombre_visible')
            fecha = request.form.get('fecha')
            categoria = request.form.get('categoria')
            notas = request.form.get('notas', '')
            
            if not impresora_id or not nombre_visible or not fecha or not categoria:
                db.close()
                return jsonify({'error': 'Faltan datos obligatorios'}), 400
                
            hoja_id = generar_id()
            extension = os.path.splitext(archivo.filename)[1] or ".pdf"
            nombre_archivo = f'hoja_{hoja_id}{extension}'
            
            archivo.save(os.path.join(ARCH_DIR, nombre_archivo))
            
            db.execute('''INSERT INTO hojas_servicio (id, impresora_id, nombre_archivo, nombre_visible, fecha, categoria, notas)
                          VALUES (?,?,?,?,?,?,?)''',
                       (hoja_id, impresora_id, nombre_archivo, nombre_visible, fecha, categoria, notas))
            db.commit()
            db.close()
            return jsonify({'id': hoja_id, 'ok': True})
        except Exception as e:
            db.close()
            return jsonify({"error": str(e)}), 500

@app.route('/api/hojas_servicio/<hoja_id>', methods=['PUT', 'DELETE'])
def handle_hoja_servicio_item(hoja_id):
    db = get_db()
    if request.method == 'PUT':
        try:
            datos = request.json
            db.execute('''UPDATE hojas_servicio 
                          SET nombre_visible=?, fecha=?, categoria=?, notas=? 
                          WHERE id=?''',
                       (datos['nombre_visible'], datos['fecha'], datos['categoria'], datos.get('notas', ''), hoja_id))
            db.commit()
            db.close()
            return jsonify({'ok': True})
        except Exception as e:
            db.close()
            return jsonify({"error": str(e)}), 500
            
    if request.method == 'DELETE':
        try:
            fila = db.execute('SELECT nombre_archivo FROM hojas_servicio WHERE id=?', (hoja_id,)).fetchone()
            if not fila:
                db.close()
                return jsonify({'error': 'Hoja de servicio no encontrada'}), 404
                
            nombre_archivo = fila['nombre_archivo']
            db.execute('DELETE FROM hojas_servicio WHERE id=?', (hoja_id,))
            db.commit()
            db.close()
            
            ruta = os.path.join(ARCH_DIR, nombre_archivo)
            if os.path.exists(ruta):
                try: os.remove(ruta)
                except: pass
                
            return jsonify({'ok': True})
        except Exception as e:
            db.close()
            return jsonify({"error": str(e)}), 500

# ── ARCHIVOS Y GENERACIÓN DE DOCX ──────────────────────────────
@app.route('/api/archivo/<tipo>/<empleado_id>', methods=['POST'])
def subir_documento(tipo, empleado_id):
    archivo = request.files.get('archivo')
    if not archivo or archivo.filename == '': 
        return jsonify({'error': 'Selecciona un documento válido'}), 400
        
    extension = os.path.splitext(archivo.filename)[1] or ".pdf"
    nombre_archivo = f'{tipo}_{empleado_id}{extension}'
    archivo.save(os.path.join(ARCH_DIR, nombre_archivo))
    
    db = get_db()
    if tipo == 'alta':
        db.execute('UPDATE empleados SET archivo_alta=? WHERE id=?', (nombre_archivo, empleado_id))
    elif tipo == 'baja':
        db.execute('UPDATE bajas SET archivo_baja=? WHERE id=?', (nombre_archivo, empleado_id))
    elif tipo == 'prestamo':
        db.execute('UPDATE prestamos SET archivo_prestamo=? WHERE id=?', (nombre_archivo, empleado_id))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'filename': nombre_archivo})

@app.route('/api/archivo/<nombre_archivo>')
def ver_documento(nombre_archivo):
    ruta = os.path.join(ARCH_DIR, nombre_archivo)
    if os.path.exists(ruta):
        return send_file(ruta)
    return jsonify({'error': 'Documento no encontrado'}), 404

@app.route('/api/docx/alta/<empleado_id>')
def docx_alta(empleado_id):
    db = get_db()
    fila = db.execute('SELECT * FROM empleados WHERE id=?', (empleado_id,)).fetchone()
    db.close()
    if not fila: 
        return jsonify({'error': 'Empleado no encontrado'}), 404
        
    empleado = dict(fila)
    cuentas = json.loads(empleado['cuentas'] or '{}')
    
    cuentas_seguras = {k: v for k, v in cuentas.items() if not k.endswith('_pass')}
    empleado['cuentas'] = cuentas_seguras
    
    ruta_docx = generar_docx_alta(empleado, obtener_equipos_de_empleado(empleado_id))
    nombre_descarga = f"Resguardo_Alta_{empleado['nombre'].replace(' ', '_')}.docx"
    return send_file(ruta_docx, as_attachment=True, download_name=nombre_descarga)

@app.route('/api/docx/baja/<baja_id>')
def docx_baja(baja_id):
    db = get_db()
    fila = db.execute('SELECT * FROM bajas WHERE id=?', (baja_id,)).fetchone()
    db.close()
    if not fila: 
        return jsonify({'error': 'Registro de baja no encontrado'}), 404
        
    baja = dict(fila)
    cuentas = json.loads(baja['cuentas'] or '{}')
    
    cuentas_seguras = {k: v for k, v in cuentas.items() if not k.endswith('_pass')}
    baja['cuentas'] = cuentas_seguras
    
    ruta_docx = generar_docx_baja(baja, json.loads(baja['equipos_snap'] or '[]'))
    nombre_descarga = f"Resguardo_Baja_{baja['nombre'].replace(' ', '_')}.docx"
    return send_file(ruta_docx, as_attachment=True, download_name=nombre_descarga)

@app.route('/api/docx/prestamo/<pid>')
def docx_prestamo(pid):
    db = get_db()
    fila = db.execute('SELECT * FROM prestamos WHERE id=?', (pid,)).fetchone()
    db.close()
    if not fila: 
        return jsonify({'error': 'Préstamo no encontrado'}), 404
        
    prestamo = dict(fila)
    ruta_docx = generar_docx_prestamo(prestamo, json.loads(prestamo['equipos'] or '[]'))
    nombre_descarga = f"Prestamo_{prestamo['nombre_solicitante'].replace(' ', '_')}.docx"
    return send_file(ruta_docx, as_attachment=True, download_name=nombre_descarga)

# ── EXPORTACIÓN A EXCEL ────────────────────────────────────────
@app.route('/api/exportar/<modulo>')
def exportar_excel(modulo):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'error': 'La librería openpyxl no está instalada. Ejecuta: pip install openpyxl'}), 500

    libro = openpyxl.Workbook()
    libro.remove(libro.active)

    estilo_fondo_cabecera = PatternFill('solid', fgColor='1F2937')
    estilo_fuente_cabecera = Font(color='FFFFFF', bold=True, size=10)
    estilo_fondo_alterno = PatternFill('solid', fgColor='F3F4F6')
    borde_fino = Side(style='thin', color='D1D5DB')
    bordes = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

    def crear_hoja(libro, titulo, cabeceras, filas):
        hoja = libro.create_sheet(title=titulo[:31])
        hoja.append(cabeceras)
        for celda in hoja[1]:
            celda.fill = estilo_fondo_cabecera
            celda.font = estilo_fuente_cabecera
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = bordes
            
        for i, fila in enumerate(filas):
            hoja.append(fila)
            fondo = estilo_fondo_alterno if i % 2 else PatternFill('solid', fgColor='FFFFFF')
            for celda in hoja[i + 2]:
                celda.fill = fondo
                celda.border = bordes
                celda.alignment = Alignment(vertical='center', wrap_text=True)
                
        for columna in hoja.columns:
            longitud_maxima = max((len(str(c.value or '')) for c in col), default=0)
            hoja.column_dimensions[columna[0].column_letter].width = min(longitud_maxima + 4, 40)
        hoja.row_dimensions[1].height = 20
        return hoja

    db = get_db()

    if modulo in ('empleados', 'todo'):
        filas = db.execute('SELECT * FROM empleados WHERE activo=1 ORDER BY nombre').fetchall()
        datos = []
        for r in filas:
            eqs = obtener_equipos_de_empleado(r['id'])
            str_equipos = '; '.join(f"{e['tipo']} {e.get('marca','')} {e.get('modelo','')}" for e in eqs)
            c = json.loads(r['cuentas'] or '{}')
            datos.append([r['nombre'], r['area'], r['direccion'], r['ubicacion'], r['telefono'],
                         r['fecha_alta'], c.get('vrtcl',''), c.get('correo',''), c.get('sap',''),
                         c.get('pos',''), c.get('req',''), str_equipos, len(eqs)])
        crear_hoja(libro, 'Empleados Activos',
            ['Nombre', 'Área', 'Dirección', 'Ubicación', 'Teléfono', 'Fecha Alta',
             'VRTCL', 'Correo', 'SAP', 'POS', 'Requisición', 'Equipos Asignados', '# Equipos'], datos)

    if modulo in ('bajas', 'todo'):
        filas = db.execute('SELECT * FROM bajas ORDER BY fecha_baja DESC').fetchall()
        datos = [[r['nombre'], r['area'], r['fecha_alta'], r['fecha_baja'], r['motivo'],
                 r['direccion'], r['ubicacion']] for r in filas]
        crear_hoja(libro, 'Bajas', ['Nombre', 'Área', 'Fecha Alta', 'Fecha Baja', 'Motivo', 'Dirección', 'Ubicación'], datos)

    if modulo in ('inventario', 'todo'):
        filas = db.execute('SELECT i.*, ip.direccion as ip_dir FROM inventario i LEFT JOIN ips ip ON i.ip_id=ip.id ORDER BY i.tipo, i.marca').fetchall()
        datos = []
        for r in filas:
            r_dict = dict(r)
            asignados = obtener_usuarios_de_equipo(r_dict['id'])
            str_asignados = '; '.join(a['nombre'] for a in asignados)
            datos.append([r_dict['tipo'], r_dict.get('categoria',''), r_dict.get('marca'), r_dict.get('modelo'), r_dict.get('num_serie'), r_dict.get('ip_dir'),
                         r_dict.get('procesador'), r_dict.get('ram'), r_dict.get('almacenamiento'),
                         r_dict.get('licencia_windows',''), r_dict.get('version_office',''), r_dict.get('contrasena',''), r_dict.get('numero_tel',''),
                         r_dict.get('cantidad_total'), r_dict.get('cantidad_libre'), r_dict.get('estado'), str_asignados, r_dict.get('notas')])
        crear_hoja(libro, 'Inventario',
            ['Tipo', 'Categoría', 'Marca', 'Modelo', 'N/S', 'IP', 'Procesador', 'RAM', 'Almacenamiento',
             'Licencia Windows', 'Office', 'Contraseña', 'Teléfono Celular', 'Total', 'Libre', 'Estado', 'Asignado a', 'Notas'], datos)

    if modulo in ('ips', 'todo'):
        filas = db.execute('SELECT * FROM ips').fetchall()
        filas_ordenadas = ordenar_ips(filas)
        datos = [[r['direccion'], r['subred'], r['estado'], r['asignada_a'], r['tipo_disp'], r['notas']] for r in filas_ordenadas]
        crear_hoja(libro, 'IPs', ['Dirección IP', 'Subred', 'Estado', 'Asignada a', 'Tipo Dispositivo', 'Notas'], datos)

    if modulo in ('impresoras', 'todo'):
        filas = db.execute('SELECT i.*, ip.direccion as ip_dir FROM impresoras i LEFT JOIN ips ip ON i.ip_id=ip.id').fetchall()
        datos = []
        for r in filas:
            eqs = db.execute('SELECT inv.tipo, inv.marca, inv.modelo, ie.tipo_conn FROM impresora_equipos ie JOIN inventario inv ON ie.inv_id=inv.id WHERE ie.imp_id=?', (r['id'],)).fetchall()
            str_equipos = '; '.join(f"{e['tipo']} {e['marca'] or ''} ({e['tipo_conn']})" for e in eqs)
            datos.append([r['nombre'], r['marca'], r['modelo'], r['tipo_consumible'], r['modelo_consumible'],
                         r['ubicacion'], r['ip_dir'], r['estado'], len(eqs), str_equipos, r['notas']])
        crear_hoja(libro, 'Impresoras',
            ['Nombre', 'Marca', 'Modelo', 'Tipo Consumible', 'Modelo Consumible',
             'Ubicación', 'IP', 'Estado', '# Equipos Conectados', 'Equipos', 'Notas'], datos)

    if modulo in ('camaras', 'todo'):
        filas = db.execute('SELECT c.*, ip.direccion as ip_dir FROM camaras c LEFT JOIN ips ip ON c.ip_id=ip.id ORDER BY c.area, c.ubicacion').fetchall()
        datos = [[r['area'], r['ubicacion'], r['marca'], r['modelo'], r['ip_dir'], r['estado'], r['notas']] for r in filas]
        crear_hoja(libro, 'Cámaras', ['Área', 'Ubicación', 'Marca', 'Modelo', 'IP', 'Estado', 'Notas'], datos)

    if modulo in ('red', 'todo'):
        filas = db.execute('SELECT r.*, ip.direccion as ip_dir FROM red_infra r LEFT JOIN ips ip ON r.ip_id=ip.id ORDER BY r.tipo').fetchall()
        datos = [[r['tipo'], r['nombre'], r['marca'], r['modelo'], r['num_serie'], r['ip_dir'],
                 r['ubicacion'], r['puertos'], r['estado'], r['notas']] for r in filas]
        crear_hoja(libro, 'Red e Infraestructura',
            ['Tipo', 'Nombre', 'Marca', 'Modelo', 'N/S', 'IP', 'Ubicación', 'Puertos', 'Estado', 'Notas'], datos)

    if modulo in ('asignaciones', 'todo'):
        filas = db.execute('''
            SELECT e.nombre, e.area, i.tipo, i.marca, i.modelo, i.num_serie, a.cantidad, a.fecha_asig
            FROM asignaciones a
            JOIN empleados e ON a.emp_id=e.id
            JOIN inventario i ON a.inv_id=i.id
            WHERE a.activa=1 AND e.activo=1
            ORDER BY e.nombre
        ''').fetchall()
        datos = [[r['nombre'], r['area'], r['tipo'], r['marca'], r['modelo'], r['num_serie'], r['cantidad'], r['fecha_asig']] for r in filas]
        crear_hoja(libro, 'Asignaciones Activas',
            ['Empleado', 'Área', 'Tipo Equipo', 'Marca', 'Modelo', 'N/S', 'Cantidad', 'Fecha Asig.'], datos)

    if modulo in ('prestamos', 'todo'):
        filas = db.execute('SELECT * FROM prestamos ORDER BY fecha_prestamo DESC').fetchall()
        datos = [[r['nombre_solicitante'], r['area'], r['fecha_prestamo'], r['fecha_esperada'], r['estado'], r['notas']] for r in filas]
        crear_hoja(libro, 'Préstamos', ['Solicitante', 'Área', 'Fecha Préstamo', 'Devolución Esperada', 'Estado', 'Notas Generales'], datos)

    db.close()

    if not libro.sheetnames: 
        libro.create_sheet('Sin datos')
        
    import tempfile
    archivo_temporal = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    libro.save(archivo_temporal.name)
    archivo_temporal.close()
    
    nombre_archivo = f"Reporte_{modulo}_{fecha_hoy()}.xlsx"
    return send_file(archivo_temporal.name, as_attachment=True, download_name=nombre_archivo)

# ── INICIO DE APLICACIÓN ───────────────────────────────────────
if __name__ == '__main__':
    init_db()
    migrar_db()
    import webbrowser, threading
    threading.Timer(1.5, lambda: webbrowser.open('http://localhost:5000')).start()
    print("\n[OK]  Resguardos Hospital Escandón v4.0 (Enterprise Edition)")
    print("   Local:  http://localhost:5000")
    print("   Red:    http://192.168.254.150:5000\n")
    app.run(host='0.0.0.0', port=5000, debug=False)